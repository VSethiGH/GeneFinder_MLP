#!/usr/bin/env python3
"""
Network Topology Analysis for C4/C6 Genesets
Compares protein-protein interactions and connectivity metrics using STRING
Adds size-matched randomization for each gene set
FIXED VERSION - Correctly handles edge vs connectivity calculations
"""

import time
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import requests


class StringNetworkAnalyzer:
    def __init__(
        self,
        excel_file,
        sheet_name,
        c4_column="G",
        c6_column="I",
        species=9606,
        required_score=150,
        n_iterations=100,
        random_seed=42,
    ):
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self.c4_column = c4_column
        self.c6_column = c6_column
        self.species = species
        self.required_score = required_score
        self.n_iterations = n_iterations
        self.random_seed = random_seed

        self.string_api_url = "https://string-db.org/api/tsv/network"
        self._network_cache = {}

    def _col_letter_to_index(self, col_letter):
        col_letter = col_letter.upper()
        index = 0
        for char in col_letter:
            index = index * 26 + (ord(char) - ord("A") + 1)
        return index - 1

    def _clean_gene(self, value):
        if pd.isna(value):
            return None
        gene = str(value).strip()
        if gene.endswith(".0"):
            gene = gene[:-2]
        gene = gene.strip()
        if not gene:
            return None
        return gene.upper()

    def load_genesets(self):
        """Load C4 and C6 genes from Excel file."""
        c4_idx = self._col_letter_to_index(self.c4_column)
        c6_idx = self._col_letter_to_index(self.c6_column)

        df = pd.read_excel(self.excel_file, sheet_name=self.sheet_name, header=None)

        c4_series = df.iloc[1:, c4_idx].dropna().map(self._clean_gene).dropna()
        c6_series = df.iloc[1:, c6_idx].dropna().map(self._clean_gene).dropna()

        self.c4_genes = set(c4_series.tolist())
        self.c6_genes = set(c6_series.tolist())

        self.intersection = self.c4_genes & self.c6_genes
        self.union = self.c4_genes | self.c6_genes
        self.only_c4 = self.c4_genes - self.c6_genes
        self.only_c6 = self.c6_genes - self.c4_genes
        self.union_minus_intersection = self.union - self.intersection

        self.background_genes = self.union

        print(f"C4 genes: {len(self.c4_genes)}")
        print(f"C6 genes: {len(self.c6_genes)}")
        print(f"Intersection: {len(self.intersection)}")
        print(f"Only C4: {len(self.only_c4)}")
        print(f"Only C6: {len(self.only_c6)}")
        print(f"Union: {len(self.union)}")
        print(f"Union - Intersection: {len(self.union_minus_intersection)}\n")

    def query_string_interactions(self, genes):
        """
        Query STRING API for protein-protein interactions.
        Returns edges and adjacency info.
        """
        genes = {self._clean_gene(g) for g in genes if self._clean_gene(g)}
        if not genes:
            return {"edges": [], "adjacency": defaultdict(set)}

        cache_key = (tuple(sorted(genes)), self.required_score, self.species)
        if cache_key in self._network_cache:
            return self._network_cache[cache_key]

        params = {
            "identifiers": "\n".join(sorted(genes)),
            "species": self.species,
            "required_score": self.required_score,
            "add_nodes": 0,
        }

        try:
            response = requests.post(self.string_api_url, data=params, timeout=60)
            response.raise_for_status()

            text = response.text.strip()
            if not text:
                result = {"edges": [], "adjacency": defaultdict(set)}
                self._network_cache[cache_key] = result
                return result

            lines = text.splitlines()
            header = lines[0].split("\t")

            gene1_idx = header.index("preferredName_A") if "preferredName_A" in header else 2
            gene2_idx = header.index("preferredName_B") if "preferredName_B" in header else 3
            score_idx = header.index("score") if "score" in header else len(header) - 1

            edges = []
            adjacency = defaultdict(set)
            seen_edges = set()

            for line in lines[1:]:
                if not line.strip():
                    continue
                parts = line.split("\t")
                if len(parts) <= max(gene1_idx, gene2_idx, score_idx):
                    continue

                try:
                    gene1 = parts[gene1_idx].strip().upper()
                    gene2 = parts[gene2_idx].strip().upper()
                    score = float(parts[score_idx])

                    if not gene1 or not gene2:
                        continue

                    edge_key = tuple(sorted((gene1, gene2)))
                    if edge_key in seen_edges:
                        continue

                    seen_edges.add(edge_key)
                    edges.append((gene1, gene2, score))
                    adjacency[gene1].add(gene2)
                    adjacency[gene2].add(gene1)

                except Exception:
                    continue

            result = {"edges": edges, "adjacency": adjacency}
            self._network_cache[cache_key] = result
            return result

        except Exception as e:
            print(f"STRING query failed: {e}")
            result = {"edges": [], "adjacency": defaultdict(set)}
            self._network_cache[cache_key] = result
            return result

    def calculate_connectivity(self, genes, interaction_data):
        """
        Connectivity = (2 * edges) / nodes
        Returns: (n_edges, connectivity)
        """
        n_nodes = len(genes)
        if n_nodes == 0:
            return 0, 0.0

        n_edges = len(interaction_data["edges"])
        connectivity = (2 * n_edges) / n_nodes
        return n_edges, connectivity

    def generate_random_geneset(self, size, background_genes, rng):
        """Generate a random gene set of specified size from background genes."""
        background_list = list(background_genes)
        if size > len(background_list):
            raise ValueError(
                f"Cannot sample {size} genes from background of size {len(background_list)}"
            )
        sampled = rng.choice(background_list, size=size, replace=False)
        return set(sampled.tolist())

    def run_randomization_test(self, size, background_genes, n_iterations=None, random_seed=None, label=""):
        """
        Run size-matched randomization test.
        Returns: (random_edges_list, random_connectivities_list)
        """
        if n_iterations is None:
            n_iterations = self.n_iterations
        if random_seed is None:
            random_seed = self.random_seed

        rng = np.random.default_rng(random_seed)

        print(f"Running randomization for {label} (size={size}, iterations={n_iterations})...")
        random_edges_list = []
        random_connectivities_list = []

        for i in range(n_iterations):
            if (i + 1) % 10 == 0:
                print(f"  Progress: {i + 1}/{n_iterations}")

            # Generate random gene set of same size as observed set
            random_genes = self.generate_random_geneset(size, background_genes, rng)
            
            # Query interactions for random set
            random_interactions = self.query_string_interactions(random_genes)
            
            # Calculate connectivity metrics
            n_edges, connectivity = self.calculate_connectivity(random_genes, random_interactions)

            random_edges_list.append(n_edges)
            random_connectivities_list.append(connectivity)

            time.sleep(0.05)

        return random_edges_list, random_connectivities_list

    def summarize_randomization(self, observed_edges, observed_connectivity, random_edges, random_connectivities):
        """
        Compare observed values to random distribution.
        Returns statistics for both edges and connectivity.
        """
        edges_arr = np.array(random_edges, dtype=float)
        conn_arr = np.array(random_connectivities, dtype=float)

        # Statistics for edges
        mean_random_edges = float(np.mean(edges_arr))
        std_random_edges = float(np.std(edges_arr))
        min_random_edges = float(np.min(edges_arr))
        max_random_edges = float(np.max(edges_arr))
        p_value_edges = float(np.mean(edges_arr >= observed_edges))  # % random >= observed
        z_score_edges = float((observed_edges - mean_random_edges) / std_random_edges) if std_random_edges > 0 else 0.0

        # Statistics for connectivity
        mean_random_conn = float(np.mean(conn_arr))
        std_random_conn = float(np.std(conn_arr))
        min_random_conn = float(np.min(conn_arr))
        max_random_conn = float(np.max(conn_arr))
        p_value_conn = float(np.mean(conn_arr >= observed_connectivity))  # % random >= observed
        z_score_conn = float((observed_connectivity - mean_random_conn) / std_random_conn) if std_random_conn > 0 else 0.0

        return {
            "observed_edges": observed_edges,
            "mean_random_edges": mean_random_edges,
            "std_random_edges": std_random_edges,
            "min_random_edges": min_random_edges,
            "max_random_edges": max_random_edges,
            "p_value_edges": p_value_edges,
            "z_score_edges": z_score_edges,
            "observed_connectivity": observed_connectivity,
            "mean_random_connectivity": mean_random_conn,
            "std_random_connectivity": std_random_conn,
            "min_random_connectivity": min_random_conn,
            "max_random_connectivity": max_random_conn,
            "p_value_connectivity": p_value_conn,
            "z_score_connectivity": z_score_conn,
        }

    def analyze_gene_set(self, name, genes, background_genes, do_random=True):
        """Analyze a single gene set."""
        genes = set(genes)

        interactions = self.query_string_interactions(genes)
        edges, connectivity = self.calculate_connectivity(genes, interactions)

        result = {
            "name": name,
            "n_genes": len(genes),
            "n_edges": edges,
            "connectivity": connectivity,
            "random_edges": None,
            "random_connectivities": None,
            "random_stats": None,
        }

        if do_random and len(genes) < len(background_genes):
            random_edges, random_connectivities = self.run_randomization_test(
                size=len(genes),
                background_genes=background_genes,
                n_iterations=self.n_iterations,
                random_seed=self.random_seed,
                label=name,
            )
            random_stats = self.summarize_randomization(
                observed_edges=edges,
                observed_connectivity=connectivity,
                random_edges=random_edges,
                random_connectivities=random_connectivities,
            )
            result["random_edges"] = random_edges
            result["random_connectivities"] = random_connectivities
            result["random_stats"] = random_stats

        return result

    def create_results_spreadsheet(self, results):
        """Create comprehensive results spreadsheet."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        # -------------------------
        # Summary sheet
        # -------------------------
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary["A1"] = "Network Topology Analysis Summary"
        ws_summary["A1"].font = Font(bold=True, size=12)

        row = 3
        headers = [
            "Metric",
            "C4 Genes",
            "C6 Genes",
            "Intersection",
            "Only C4",
            "Only C6",
            "Union",
            "Union - Intersection",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row += 1
        ws_summary.cell(row=row, column=1, value="Number of Genes")
        ws_summary.cell(row=row, column=2, value=results["C4 Genes"]["n_genes"])
        ws_summary.cell(row=row, column=3, value=results["C6 Genes"]["n_genes"])
        ws_summary.cell(row=row, column=4, value=results["Intersection"]["n_genes"])
        ws_summary.cell(row=row, column=5, value=results["Only C4"]["n_genes"])
        ws_summary.cell(row=row, column=6, value=results["Only C6"]["n_genes"])
        ws_summary.cell(row=row, column=7, value=results["Union"]["n_genes"])
        ws_summary.cell(row=row, column=8, value=results["Union - Intersection"]["n_genes"])

        row += 1
        ws_summary.cell(row=row, column=1, value="Number of Edges (PPI)")
        ws_summary.cell(row=row, column=2, value=results["C4 Genes"]["n_edges"])
        ws_summary.cell(row=row, column=3, value=results["C6 Genes"]["n_edges"])
        ws_summary.cell(row=row, column=4, value=results["Intersection"]["n_edges"])
        ws_summary.cell(row=row, column=5, value=results["Only C4"]["n_edges"])
        ws_summary.cell(row=row, column=6, value=results["Only C6"]["n_edges"])
        ws_summary.cell(row=row, column=7, value=results["Union"]["n_edges"])
        ws_summary.cell(row=row, column=8, value=results["Union - Intersection"]["n_edges"])

        row += 1
        ws_summary.cell(row=row, column=1, value="Connectivity (2*edges/nodes)")
        ws_summary.cell(row=row, column=2, value=f'{results["C4 Genes"]["connectivity"]:.4f}')
        ws_summary.cell(row=row, column=3, value=f'{results["C6 Genes"]["connectivity"]:.4f}')
        ws_summary.cell(row=row, column=4, value=f'{results["Intersection"]["connectivity"]:.4f}')
        ws_summary.cell(row=row, column=5, value=f'{results["Only C4"]["connectivity"]:.4f}')
        ws_summary.cell(row=row, column=6, value=f'{results["Only C6"]["connectivity"]:.4f}')
        ws_summary.cell(row=row, column=7, value=f'{results["Union"]["connectivity"]:.4f}')
        ws_summary.cell(row=row, column=8, value=f'{results["Union - Intersection"]["connectivity"]:.4f}')

        # -------------------------
        # Randomization summary sheet
        # -------------------------
        ws_rand = wb.create_sheet("Randomization Summary")
        ws_rand["A1"] = "Randomization Summary (vs Random Gene Sets)"
        ws_rand["A1"].font = Font(bold=True, size=12)

        rand_row = 3
        rand_headers = [
            "Gene Set",
            "Size",
            "Observed Edges",
            "Mean Random Edges",
            "Std Dev Random Edges",
            "Z-score Edges",
            "P-value Edges",
            "Observed Connectivity",
            "Mean Random Connectivity",
            "Std Dev Random Connectivity",
            "Z-score Connectivity",
            "P-value Connectivity",
        ]
        for col_idx, header in enumerate(rand_headers, start=1):
            cell = ws_rand.cell(row=rand_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        rand_row += 1
        for set_name in ["C4 Genes", "C6 Genes", "Intersection", "Only C4", "Only C6", "Union - Intersection"]:
            r = results[set_name]
            if r["random_stats"] is None:
                continue
            s = r["random_stats"]

            ws_rand.cell(row=rand_row, column=1, value=set_name)
            ws_rand.cell(row=rand_row, column=2, value=r["n_genes"])
            ws_rand.cell(row=rand_row, column=3, value=s["observed_edges"])
            ws_rand.cell(row=rand_row, column=4, value=f'{s["mean_random_edges"]:.4f}')
            ws_rand.cell(row=rand_row, column=5, value=f'{s["std_random_edges"]:.4f}')
            ws_rand.cell(row=rand_row, column=6, value=f'{s["z_score_edges"]:.4f}')
            ws_rand.cell(row=rand_row, column=7, value=f'{s["p_value_edges"]:.4f}')
            ws_rand.cell(row=rand_row, column=8, value=f'{s["observed_connectivity"]:.4f}')
            ws_rand.cell(row=rand_row, column=9, value=f'{s["mean_random_connectivity"]:.4f}')
            ws_rand.cell(row=rand_row, column=10, value=f'{s["std_random_connectivity"]:.4f}')
            ws_rand.cell(row=rand_row, column=11, value=f'{s["z_score_connectivity"]:.4f}')
            ws_rand.cell(row=rand_row, column=12, value=f'{s["p_value_connectivity"]:.4f}')
            rand_row += 1

        # -------------------------
        # Random distributions sheet
        # -------------------------
        ws_dist = wb.create_sheet("Random Distributions")
        ws_dist["A1"] = "Iteration"
        ws_dist["A1"].font = Font(bold=True)

        dist_sets = ["C4 Genes", "C6 Genes", "Intersection", "Only C4", "Only C6", "Union - Intersection"]
        for col_idx, set_name in enumerate(dist_sets, start=2):
            cell = ws_dist.cell(row=1, column=col_idx, value=f"{set_name} Connectivity")
            cell.font = Font(bold=True)

        for i in range(self.n_iterations):
            ws_dist.cell(row=i + 2, column=1, value=i + 1)
            for col_idx, set_name in enumerate(dist_sets, start=2):
                if results[set_name]["random_connectivities"] is not None:
                    values = results[set_name]["random_connectivities"]
                    if i < len(values):
                        ws_dist.cell(row=i + 2, column=col_idx, value=f"{values[i]:.4f}")

        # -------------------------
        # Gene sets sheet
        # -------------------------
        ws_genes = wb.create_sheet("Gene Sets")
        gene_sets_order = ["C4 Genes", "C6 Genes", "Intersection", "Only C4", "Only C6", "Union - Intersection", "Union"]
        for col_idx, set_name in enumerate(gene_sets_order, start=1):
            cell = ws_genes.cell(row=1, column=col_idx, value=set_name)
            cell.font = Font(bold=True)

        max_len = max(len(results[name].get("genes", [])) for name in gene_sets_order if name in results)

        for row_idx in range(max_len):
            for col_idx, set_name in enumerate(gene_sets_order, start=1):
                genes_list = results[set_name].get("genes", [])
                if row_idx < len(genes_list):
                    ws_genes.cell(row=row_idx + 2, column=col_idx, value=genes_list[row_idx])

        # Column widths
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws_summary.column_dimensions[col].width = 22

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws_rand.column_dimensions[col].width = 20

        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws_dist.column_dimensions[col].width = 24

        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws_genes.column_dimensions[col].width = 18

        output_file = Path(self.excel_file).stem + "_RESULTS.xlsx"
        wb.save(output_file)
        print(f"\nResults saved to: {output_file}")
        return output_file


def main():
    EXCEL_FILE = input("Enter path to your Excel file: ").strip()
    SHEET_NAME = input("Enter sheet name: ").strip()
    C4_COL = input("Enter column letter for C4 genes (default 'G'): ").strip() or "G"
    C6_COL = input("Enter column letter for C6 genes (default 'I'): ").strip() or "I"
    N_ITERATIONS = 100

    analyzer = StringNetworkAnalyzer(
        EXCEL_FILE,
        sheet_name=SHEET_NAME,
        c4_column=C4_COL,
        c6_column=C6_COL,
        species=9606,
        required_score=150,
        n_iterations=N_ITERATIONS,
        random_seed=42,
    )

    print("\n=== NETWORK TOPOLOGY ANALYSIS ===\n")
    analyzer.load_genesets()

    results = {}

    gene_set_map = {
        "C4 Genes": analyzer.c4_genes,
        "C6 Genes": analyzer.c6_genes,
        "Intersection": analyzer.intersection,
        "Only C4": analyzer.only_c4,
        "Only C6": analyzer.only_c6,
        "Union - Intersection": analyzer.union_minus_intersection,
        "Union": analyzer.union,
    }

    for set_name, genes in gene_set_map.items():
        print(f"\nAnalyzing {set_name}...")
        interactions = analyzer.query_string_interactions(genes)
        edges, connectivity = analyzer.calculate_connectivity(genes, interactions)

        do_random = set_name != "Union" and len(genes) < len(analyzer.background_genes)
        random_edges = None
        random_connectivities = None
        random_stats = None

        if do_random:
            random_edges, random_connectivities = analyzer.run_randomization_test(
                size=len(genes),
                background_genes=analyzer.background_genes,
                n_iterations=N_ITERATIONS,
                random_seed=42,
                label=set_name,
            )
            random_stats = analyzer.summarize_randomization(
                observed_edges=edges,
                observed_connectivity=connectivity,
                random_edges=random_edges,
                random_connectivities=random_connectivities,
            )

        results[set_name] = {
            "genes": sorted(list(genes)),
            "n_genes": len(genes),
            "n_edges": edges,
            "connectivity": connectivity,
            "random_edges": random_edges,
            "random_connectivities": random_connectivities,
            "random_stats": random_stats,
        }

        print(f"  Genes: {len(genes)}")
        print(f"  Edges: {edges}")
        print(f"  Connectivity: {connectivity:.4f}")
        if random_stats is not None:
            print(f"  Random mean edges: {random_stats['mean_random_edges']:.2f} ± {random_stats['std_random_edges']:.2f}")
            print(f"  Random mean connectivity: {random_stats['mean_random_connectivity']:.4f} ± {random_stats['std_random_connectivity']:.4f}")
            print(f"  Z-score connectivity: {random_stats['z_score_connectivity']:.2f}")
            print(f"  P-value connectivity: {random_stats['p_value_connectivity']:.4f}")

    print("\n=== RESULTS SUMMARY ===")
    for set_name in ["C4 Genes", "C6 Genes", "Intersection", "Only C4", "Only C6", "Union - Intersection", "Union"]:
        r = results[set_name]
        print(f"{set_name}: {r['connectivity']:.4f} (edges={r['n_edges']}, genes={r['n_genes']})")

    results_file = analyzer.create_results_spreadsheet(results)
    print(f"\n✓ Analysis complete!")
    print(f"  Results spreadsheet: {results_file}")


if __name__ == "__main__":
    main()